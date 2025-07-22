import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import json
from openpyxl import load_workbook
from datetime import datetime

CONFIG_FILE = os.path.expanduser('~/.txt_processor_config.json')

def load_config():
    try:
        with open(CONFIG_FILE) as f:
            return json.load(f)
    except:
        return {'input_folder': '', 'excel_path': '', 'file_encoding': 'auto'}

def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f)

class Application(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.config = load_config()
        self.sheet_options = ['DY2011', 'DY2012', 'DY2021', 'DY2022', 'DY2023', 'DY2024', 'DY2025']
        self.encoding_options = ['auto', 'utf-8', 'gbk', 'gb18030', 'big5']
        self.master = master
        self.pack()
        self.create_widgets()

    def create_widgets(self):
        # 输入文件夹选择
        self.input_btn = tk.Button(self, text='选择输入文件夹', command=self.select_input)
        self.input_btn.grid(row=0, column=0)
        self.input_label = tk.Label(self, text=self.config.get('input_folder', ''))
        self.input_label.grid(row=0, column=1)

        # Excel文件选择
        self.excel_btn = tk.Button(self, text='选择Excel文件', command=self.select_excel)
        self.excel_btn.grid(row=1, column=0)
        self.excel_label = tk.Label(self, text=self.config.get('excel_path', ''))
        self.excel_label.grid(row=1, column=1)

        # 工作表选择
        self.sheet_label = tk.Label(self, text='选择工作表:')
        self.sheet_label.grid(row=2, column=0)
        self.sheet_combo = ttk.Combobox(self, values=self.sheet_options)
        self.sheet_combo.set(self.config.get('sheet_name', 'DY2011'))
        self.sheet_combo.grid(row=2, column=1)

        # 编码选择
        self.encoding_label = tk.Label(self, text='文件编码:')
        self.encoding_label.grid(row=3, column=0)
        self.encoding_combo = ttk.Combobox(self, values=self.encoding_options)
        self.encoding_combo.set(self.config.get('file_encoding', 'auto'))
        self.encoding_combo.grid(row=3, column=1)

        # 进度条
        self.progress = ttk.Progressbar(self, orient='horizontal', length=200, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=2, pady=5)

        self.run_btn = tk.Button(self, text='开始处理', command=self.process_files)
        self.run_btn.grid(row=6, column=0, columnspan=2)

    def select_input(self):
        path = filedialog.askdirectory(initialdir=self.config.get('input_folder'))
        if path:
            self.config['input_folder'] = path
            self.input_label.config(text=path)
            save_config(self.config)

    def select_excel(self):
        path = filedialog.askopenfilename(
            filetypes=[('Excel文件', '*.xlsx')],
            initialdir=os.path.dirname(self.config.get('excel_path', ''))
        )
        if path:
            self.config['excel_path'] = path
            self.excel_label.config(text=path)
            save_config(self.config)

    def extract_order_numbers(self, lines):
        """从文件内容中提取订单号"""
        order_numbers = []
        
        # 找到PlateA Assignments和Protocol Path之间的内容
        start_idx = None
        end_idx = None
        
        for i, line in enumerate(lines):
            if 'PlateA Assignments' in line:
                start_idx = i + 1  # 跳过标题行
            elif 'Protocol Path=' in line:
                end_idx = i
                break
        
        if start_idx is None or end_idx is None:
            return order_numbers
        
        # 解析订单号
        for line in lines[start_idx:end_idx]:
            line = line.strip()
            if not line or '\t' not in line:
                continue
            
            parts = line.split('\t')
            if len(parts) >= 2:
                # 提取第二列的内容，如S2507080005-0001
                order_part = parts[1]
                
                # 提取基础订单号（去掉-或_后面的所有内容）
                if '-' in order_part:
                    order_num = order_part.split('-')[0]
                elif '_' in order_part:
                    order_num = order_part.split('_')[0]
                else:
                    order_num = order_part
                
                if order_num and order_num not in order_numbers:
                    order_numbers.append(order_num)
        
        return order_numbers

    def process_files(self):
        try:
            wb = load_workbook(self.config['excel_path'])
            sheet_name = self.sheet_combo.get()
            self.config['sheet_name'] = sheet_name
            save_config(self.config)
            
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(['文件名', '最后更新时间'])  # 创建新表时添加标题
            else:
                ws = wb[sheet_name]
            
            # 获取现有文件名
            existing_files = {row[0]: (row[1], row[2] if len(row) > 2 else '') for row in ws.iter_rows(min_row=2, max_col=3, values_only=True) if row[0]}
            # 添加标题行（如果不存在）
            if ws.max_row == 0 or ws['A1'].value != '文件名':
                ws.append(['文件名', '最后更新时间', '订单号'])

            file_list = [f for f in os.listdir(self.config['input_folder']) if f.endswith('.txt')]
            total_files = len(file_list)
            self.progress['maximum'] = total_files
            
            for i, filename in enumerate(file_list):
                filepath = os.path.join(self.config['input_folder'], filename)
                selected_encoding = self.config.get('file_encoding', 'auto')
                if selected_encoding == 'auto':
                    import chardet
                    with open(filepath, 'rb') as f:
                        rawdata = f.read(10000)
                        result = chardet.detect(rawdata)
                        detected_encoding = result['encoding'] or 'utf-8'
                    encoding = detected_encoding
                else:
                    encoding = selected_encoding

                try:
                    with open(filepath, 'r', encoding=encoding, errors='replace') as f:
                        lines = f.readlines()
                        if not lines:
                            print(f'跳过空文件 {filename}')
                            continue
                        last_line = lines[-1].strip()
                        
                        # 根据示例文件格式处理（空格分隔）
                        fields = last_line.split()
                        if len(fields) < 2:
                            print(f'跳过格式不正确的文件 {filename}')
                            continue
                        timestamp = ' '.join(fields[0:2])  # 提取日期和时间
                        
                        # 提取订单号
                        try:
                            order_numbers = self.extract_order_numbers(lines)
                            orders_str = '/'.join(order_numbers) if order_numbers else ''
                        except Exception as e:
                            print(f'跳过提取订单号失败的文件 {filename}: {str(e)}')
                            continue
                        
                        # 检查是否需要更新或添加记录
                        if filename in existing_files:
                            existing_timestamp, existing_orders = existing_files[filename]
                            needs_update = False
                            
                            if existing_timestamp != timestamp:
                                needs_update = True
                            if existing_orders != orders_str:
                                needs_update = True
                            
                            if needs_update:
                                # 找到对应行进行更新
                                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
                                    if row[0] == filename:
                                        ws.cell(row=row_idx, column=2, value=timestamp)
                                        ws.cell(row=row_idx, column=3, value=orders_str)
                                        break
                            continue  # 跳过已存在的文件但数据未变的情况
                        else:
                            ws.append([filename, timestamp, orders_str])
                except UnicodeDecodeError:
                    print(f'跳过编码错误的文件 {filename}')
                    continue
                except Exception as e:
                    print(f'跳过无法处理的文件 {filename}: {str(e)}')
                    continue
                finally:
                    self.progress['value'] = i+1
                self.progress.update_idletasks()  # 强制刷新进度条

            wb.save(self.config['excel_path'])
            messagebox.showinfo('完成', f'成功处理{ws.max_row-1}条记录')
        except Exception as e:
            error_msg = f'处理失败：{str(e)}\n\n错误文件：{filename if "filename" in locals() else "未知"}'
            messagebox.showerror('错误', error_msg)

if __name__ == '__main__':
    root = tk.Tk()
    root.title('TXT文件处理器')
    app = Application(master=root)
    app.mainloop()